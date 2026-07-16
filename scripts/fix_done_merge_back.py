import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Font

BLACK = "FF000000"
RED = "FFFF0000"

BLACK_FONT = InlineFont(color=BLACK)
RED_FONT = InlineFont(color=RED)


def build_richtext(label, value):
    """Build single-line rich text: label in black + value in red."""
    if value is None:
        return CellRichText(TextBlock(BLACK_FONT, str(label)))
    return CellRichText(
        TextBlock(BLACK_FONT, str(label)),
        TextBlock(RED_FONT, str(value)),
    )


def fix_canada(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    # Bank info rows
    for row in range(29, 37):
        b = ws.cell(row=row, column=2)
        c = ws.cell(row=row, column=3)
        if b.value is None:
            continue
        label = str(b.value).strip()
        value = str(c.value).strip() if c.value else ""
        # Rebuild as single-cell rich text in column B
        b.value = build_richtext(label, value)
        c.value = None

    wb.save(path)
    print(f'Fixed Canada main merged back: {path}')


def fix_usa(path):
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    # Read all label/value pairs from B11-B16 and C11-C16
    lines = []
    for row in range(11, 17):
        b = ws.cell(row=row, column=2)
        c = ws.cell(row=row, column=3)
        if b.value is None and c.value is None:
            continue
        label = str(b.value).strip() if b.value else ""
        value = str(c.value).strip() if c.value else ""
        lines.append((label, value))

    if not lines:
        print(f'No bank info found in {path}')
        return

    # Build multiline rich text: each line is label (black) + value (red), lines joined by \n
    # Actually each newline between lines should also be black (or neutral)
    rich_parts = []
    for i, (label, value) in enumerate(lines):
        if i > 0:
            rich_parts.append(TextBlock(BLACK_FONT, "\n"))
        rich_parts.append(TextBlock(BLACK_FONT, str(label)))
        rich_parts.append(TextBlock(RED_FONT, str(value)))

    # Set A11
    a11 = ws.cell(row=11, column=1)
    a11.value = CellRichText(*rich_parts)

    # Clear B/C 11
    ws.cell(row=11, column=2).value = None
    ws.cell(row=11, column=3).value = None

    # Delete extra rows 12-16 (5 rows)
    ws.delete_rows(12, 5)

    # Restore merges: D10:D11, E10:E11, A11:C11
    # Check existing merges first to avoid duplicates
    existing = set(str(r) for r in ws.merged_cells.ranges)

    def safe_merge(start, end):
        rng = f"{start}:{end}"
        if rng not in existing:
            ws.merge_cells(f"{start}:{end}")

    # After deleting 5 rows, old row 11 is still row 11
    safe_merge("D10", "D11")
    safe_merge("E10", "E11")
    safe_merge("A11", "C11")

    wb.save(path)
    print(f'Fixed USA main merged back: {path}')


# Run
fix_canada('发票模版-评审/done/FBU-加拿大-5%GST-负数.xlsx')
fix_canada('发票模版-评审/done/FBU-加拿大-5%GST.xlsx')
fix_usa('发票模版-评审/done/FBU-美国EL发票模板-US credit note（负数）.xlsx')
fix_usa('发票模版-评审/done/FBU-美国GA发票模板-GAUS credit note（负.xlsx')

print('All done.')
