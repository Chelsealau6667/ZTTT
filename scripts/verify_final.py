import openpyxl
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

def check_annotations(filepath, label, cols):
    """Check that annotation columns are cleared"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    issues = []
    for row in range(1, min(ws.max_row + 1, 40)):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            if cell.value:
                issues.append(f"  [{row},{col}] still has value: '{cell.value}'")
    if issues:
        print(f"\n{label} - ISSUES:")
        for issue in issues:
            print(issue)
    else:
        print(f"{label} - Annotation columns cleared ✓")

check_annotations(os.path.join(REVIEW_DIR, "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx"), "Canada HST Pos", [6,7,8])
check_annotations(os.path.join(REVIEW_DIR, "FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx"), "Vancouver HST Pos", [6,7,8])
check_annotations(os.path.join(REVIEW_DIR, "FBU-澳洲谷仓开票模板.xlsx"), "Australia", [7,8])
check_annotations(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"), "Japan", [7,8])
check_annotations(os.path.join(REVIEW_DIR, "ABU-深圳云颂开票模板.xlsx"), "Shenzhen", [7,8])
check_annotations(os.path.join(REVIEW_DIR, "ABU-香港云颂开票模板.xlsx"), "Hong Kong", [7,8])
