import openpyxl
from openpyxl.styles import Font
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

# Color constants
BLACK = Font(color="FF000000")
RED_MAIN = Font(color="FFFF0000")
RED_RULE = Font(color="FFF54A45")

def copy_style_base(font_color):
    return Font(color=font_color)

def fix_vancouver(filepath):
    """Insert Bank Transit + Financial Institution into Vancouver 4 files"""
    wb = openpyxl.load_workbook(filepath)
    ws_main = wb.worksheets[0]
    ws_rule = wb.worksheets[1]
    
    # ===== Main template sheet =====
    # Insert 2 rows before row 35 (which currently has SWIFT Code)
    ws_main.insert_rows(35, amount=2)
    
    # Row 35: Bank Transit
    b35 = ws_main.cell(row=35, column=2)
    b35.value = "Bank Transit:"
    b35.font = Font(name=b35.font.name if b35.font else None, size=b35.font.size if b35.font else None, color="FF000000")
    
    c35 = ws_main.cell(row=35, column=3)
    c35.value = "Bank Transit"
    c35.font = Font(name=c35.font.name if c35.font else None, size=c35.font.size if c35.font else None, color="FFFF0000")
    
    # Row 36: Fiancial Institution
    b36 = ws_main.cell(row=36, column=2)
    b36.value = "Fiancial Institution:"
    b36.font = Font(name=b36.font.name if b36.font else None, size=b36.font.size if b36.font else None, color="FF000000")
    
    c36 = ws_main.cell(row=36, column=3)
    c36.value = "Fiancial Institution"
    c36.font = Font(name=c36.font.name if c36.font else None, size=c36.font.size if c36.font else None, color="FFFF0000")
    
    # ===== Field rules sheet =====
    # Insert 2 rows before row 28 (which currently has 银行联行码)
    ws_rule.insert_rows(28, amount=2)
    
    # Row 28: Bank Transit
    for col, val in [(2, "Bank Transit"), (3, "bank_transit"), (4, "销方/收款信息"), (5, "字段映射"), (6, "FSSC.bank_transit"), (8, "可配置化")]:
        cell = ws_rule.cell(row=28, column=col)
        cell.value = val
        cell.font = Font(name=cell.font.name if cell.font else None, size=cell.font.size if cell.font else None, color="FFF54A45")
    
    # Row 29: Fiancial Institution
    for col, val in [(2, "Fiancial Institution"), (3, "fiancial_institution"), (4, "销方/收款信息"), (5, "字段映射"), (6, "FSSC.fiancial_institution"), (8, "可配置化")]:
        cell = ws_rule.cell(row=29, column=col)
        cell.value = val
        cell.font = Font(name=cell.font.name if cell.font else None, size=cell.font.size if cell.font else None, color="FFF54A45")
    
    wb.save(filepath)
    print(f"[FIXED Vancouver] {os.path.basename(filepath)}")


def fix_japan(filepath):
    """Restructure お振込先 block in Japan template"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Row 41: 银行名称 → B=label, C=value
    a41 = ws.cell(row=41, column=1)
    a41.value = None
    a41.font = Font(color="FF000000")
    
    b41 = ws.cell(row=41, column=2)
    b41.value = "银行名称:"
    b41.font = Font(name=b41.font.name if b41.font else None, size=b41.font.size if b41.font else None, color="FF000000")
    
    c41 = ws.cell(row=41, column=3)
    c41.value = "银行名称"
    c41.font = Font(name=c41.font.name if c41.font else None, size=c41.font.size if c41.font else None, color="FFFF0000")
    
    # Row 42: 账户名称 → B=label, C=value
    a42 = ws.cell(row=42, column=1)
    a42.value = None
    a42.font = Font(color="FF000000")
    
    b42 = ws.cell(row=42, column=2)
    b42.value = "账户名称:"
    b42.font = Font(name=b42.font.name if b42.font else None, size=b42.font.size if b42.font else None, color="FF000000")
    
    c42 = ws.cell(row=42, column=3)
    c42.value = "账户名称"
    c42.font = Font(name=c42.font.name if c42.font else None, size=c42.font.size if c42.font else None, color="FFFF0000")
    
    wb.save(filepath)
    print(f"[FIXED Japan] {os.path.basename(filepath)}")


def fix_yunsong(filepath):
    """Strip trailing Chinese value names from B-column labels in Yunsong templates"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    replacements = {
        19: (" 收款人名称：\n (Name Of Beneficiary)", "销售方名称"),
        20: (" 银行名称\n (Bank Name)", "银行名称"),
        21: (" 账户名称\n Account Name", "账户名称"),
        22: (" 账户号\n (Account Number)", "账户号"),
        23: (" 银行代码\n (Bank code)", "银行代码"),
        24: (" 银行联行码\n (Bank Swift Code)", "银行联行码"),
        25: (" 银行地址\n (Bank Address)", "银行地址"),
        26: (" 发票备注：\n (Comment)", None),
    }
    
    for row, (new_label, c_value) in replacements.items():
        b = ws.cell(row=row, column=2)
        # Keep original font style but ensure black
        orig_font = b.font
        b.value = new_label
        b.font = Font(name=orig_font.name, size=orig_font.size, bold=orig_font.bold,
                      italic=orig_font.italic, color="FF000000")
        # Note: we don't set C value because it's already correct and red
    
    wb.save(filepath)
    print(f"[FIXED Yunsong] {os.path.basename(filepath)}")


def main():
    vancouver_files = [
        "FBU-温哥华模版-GST税率(0%、5%）正数.xlsx",
        "FBU-温哥华模版-GST税率(0%、5%）负数.xlsx",
        "FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx",
        "FBU-温哥华模版-HST税率（13%、14%、15%）负数.xlsx",
    ]
    for f in vancouver_files:
        fix_vancouver(os.path.join(REVIEW_DIR, f))
    
    fix_japan(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"))
    
    for f in ["ABU-深圳云颂开票模板.xlsx", "ABU-香港云颂开票模板.xlsx"]:
        fix_yunsong(os.path.join(REVIEW_DIR, f))
    
    print("\nAll fixes applied.")

if __name__ == "__main__":
    main()
