import openpyxl
from openpyxl.styles import Font

RED = "FFFF0000"
BLACK = "FF000000"
BLACK_FONT = Font(color=BLACK)
RED_FONT = Font(color=RED)


def fix_usa_main(ws):
    a11 = ws.cell(row=11, column=1)
    if not a11.value:
        return
    lines = str(a11.value).strip().split("\n")
    entries = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Use full-width colon first (value separator in these files)
        idx = line.rfind("：")
        if idx == -1:
            # fallback to English colon+space
            idx = line.rfind(": ")
        if idx != -1:
            label = line[:idx + 1]
            value = line[idx + 1:]
            entries.append((label, value))
        else:
            entries.append((line, ""))

    # Unmerge ranges that involve row 11 bank block and D10/E10 merges
    ranges_to_unmerge = []
    for mc_range in ws.merged_cells.ranges:
        if 'A11' in str(mc_range) or 'B11' in str(mc_range) or 'C11' in str(mc_range):
            ranges_to_unmerge.append(str(mc_range))
        if 'D10' in str(mc_range) or 'D11' in str(mc_range):
            ranges_to_unmerge.append(str(mc_range))
        if 'E10' in str(mc_range) or 'E11' in str(mc_range):
            ranges_to_unmerge.append(str(mc_range))

    for rng in set(ranges_to_unmerge):
        ws.unmerge_cells(rng)

    # Clear A11
    a11.value = None

    # First entry at row 11 B/C
    b = ws.cell(row=11, column=2)
    c = ws.cell(row=11, column=3)
    b.value = entries[0][0]
    b.font = BLACK_FONT
    c.value = entries[0][1]
    c.font = RED_FONT

    # Remaining entries: insert rows at 12, 13...
    for i in range(1, len(entries)):
        ws.insert_rows(11 + i, amount=1)
        b = ws.cell(row=11 + i, column=2)
        c = ws.cell(row=11 + i, column=3)
        b.value = entries[i][0]
        b.font = BLACK_FONT
        c.value = entries[i][1]
        c.font = RED_FONT


# Run for USA files
for f in ['发票模版-评审/done/FBU-美国EL发票模板-US credit note（负数）.xlsx',
          '发票模版-评审/done/FBU-美国GA发票模板-GAUS credit note（负.xlsx']:
    wb = openpyxl.load_workbook(f)
    fix_usa_main(wb[wb.sheetnames[0]])
    wb.save(f)
    print(f'Fixed USA main: {f}')

print('Done.')
