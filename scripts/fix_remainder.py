import openpyxl
from openpyxl.styles import Font
import os
from copy import copy

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

def set_font_color(cell, color_code):
    f = copy(cell.font) if cell.font else Font()
    f.color.rgb = color_code
    cell.font = f

def fix_japan(filepath):
    """Restructure お振込先 block: unmerge A:B, set labels in B, values in C"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Unmerge A41:B41 and A42:B42
    for mr in list(ws.merged_cells.ranges):
        if (mr.min_row == 41 and mr.max_row == 41 and mr.min_col == 1 and mr.max_col == 2) or \
           (mr.min_row == 42 and mr.max_row == 42 and mr.min_col == 1 and mr.max_col == 2):
            ws.unmerge_cells(str(mr))
    
    # Row 41: 银行名称
    a41 = ws.cell(row=41, column=1)
    a41.value = None
    set_font_color(a41, "FF000000")
    
    b41 = ws.cell(row=41, column=2)
    b41.value = "银行名称:"
    set_font_color(b41, "FF000000")
    
    c41 = ws.cell(row=41, column=3)
    c41.value = "银行名称"
    set_font_color(c41, "FFFF0000")
    
    # Row 42: 账户名称
    a42 = ws.cell(row=42, column=1)
    a42.value = None
    set_font_color(a42, "FF000000")
    
    b42 = ws.cell(row=42, column=2)
    b42.value = "账户名称:"
    set_font_color(b42, "FF000000")
    
    c42 = ws.cell(row=42, column=3)
    c42.value = "账户名称"
    set_font_color(c42, "FFFF0000")
    
    wb.save(filepath)
    print(f"[FIXED Japan main] {os.path.basename(filepath)}")


def fix_yunsong(filepath):
    """Strip trailing Chinese value names from B-column labels"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # These replacements have the same value currently; just need to strip trailing Chinese
    # from B column. The merged ranges don't affect B in rows 19-26.
    replacements = {
        19: " 收款人名称：\n (Name Of Beneficiary)",
        20: " 银行名称\n (Bank Name)",
        21: " 账户名称\n Account Name",
        22: " 账户号\n (Account Number)",
        23: " 银行代码\n (Bank code)",
        24: " 银行联行码\n (Bank Swift Code)",
        25: " 银行地址\n (Bank Address)",
        26: " 发票备注：\n (Comment)",
    }
    
    for row, new_label in replacements.items():
        b = ws.cell(row=row, column=2)
        b.value = new_label
        set_font_color(b, "FF000000")
    
    wb.save(filepath)
    print(f"[FIXED Yunsong] {os.path.basename(filepath)}")


def main():
    fix_japan(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"))
    for f in ["ABU-深圳云颂开票模板.xlsx", "ABU-香港云颂开票模板.xlsx"]:
        fix_yunsong(os.path.join(REVIEW_DIR, f))
    print("\nRemaining fixes applied.")

if __name__ == "__main__":
    main()
