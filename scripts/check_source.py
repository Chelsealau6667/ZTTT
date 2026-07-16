import openpyxl
import os

SOURCE = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审/形式发票模板合并 (5).xlsx"
wb = openpyxl.load_workbook(SOURCE)

# Find the Canada sheet
for name in wb.sheetnames:
    if "加拿大" in name or "Canada" in name:
        ws = wb[name]
        print(f"Sheet: {name}")
        # Print rows 25-40, cols A-H
        for row in range(25, min(41, ws.max_row + 1)):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    color = None
                    if cell.font and cell.font.color:
                        color = cell.font.color.rgb
                    print(f"  [{row},{col}] '{cell.value}' color={color}")
        print("---")
