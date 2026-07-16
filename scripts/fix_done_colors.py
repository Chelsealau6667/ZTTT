import openpyxl
from openpyxl.styles import Font
from copy import copy

RED = "FFFF0000"
BLACK = "FF000000"
BLACK_FONT = Font(color=BLACK)
RED_FONT = Font(color=RED)

def fix_canada_main(ws):
    rows_to_fix = [29, 30, 31, 32, 33, 34, 35, 36]
    for r in rows_to_fix:
        b = ws.cell(row=r, column=2)
        if not b.value:
            continue
        text = str(b.value).strip()
        idx = -1
        sep_used = None
        for sep in ["：", ":（", ":"]:
            i = text.find(sep)
            if i != -1:
                idx = i
                sep_used = sep
                break
        if idx == -1:
            continue
        label = text[:idx + len(sep_used)]
        value = text[idx + len(sep_used):].strip()
        # Fix Account# value: replace (收款币种) with (币种)
        if "Account#" in label:
            value = value.replace("（收款币种）", "（币种）")
        b.value = label
        b.font = BLACK_FONT
        c = ws.cell(row=r, column=3)
        c.value = value
        c.font = RED_FONT

def fix_usa_main(ws):
    a11 = ws.cell(row=11, column=1)
    if not a11.value:
        return
    lines = str(a11.value).strip().split("\n")
    entries = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Try full-width colon first, then colon+space
        idx = line.rfind("：")
        if idx == -1:
            idx = line.rfind(": ")
        if idx != -1:
            label = line[:idx + 1]
            value = line[idx + 1:]
            entries.append((label, value))
        else:
            entries.append((line, ""))
    a11.value = None
    # First entry at row 11 B/C
    b = ws.cell(row=11, column=2)
    c = ws.cell(row=11, column=3)
    b.value = entries[0][0]
    b.font = BLACK_FONT
    c.value = entries[0][1]
    c.font = RED_FONT
    for i in range(1, len(entries)):
        ws.insert_rows(11 + i, amount=1)
        b = ws.cell(row=11 + i, column=2)
        c = ws.cell(row=11 + i, column=3)
        b.value = entries[i][0]
        b.font = BLACK_FONT
        c.value = entries[i][1]
        c.font = RED_FONT

def fix_rules_structure(path):
    wb = openpyxl.load_workbook(path)
    ws = wb['字段取值规则']
    # Check H3; if None, the original file didn't have H column
    h3 = ws.cell(row=3, column=8).value
    # Fix Canada 5% GST where G3 might have been overwritten
    # Ensure headers match Vancouver: G=日期格式, H=解析类型
    # If G3=='解析类型' and H3 is None -> insert column at G and set G='日期格式'
    g3 = ws.cell(row=3, column=7).value
    if g3 == '解析类型' and h3 is None:
        # Insert a new column G for 日期格式
        ws.insert_cols(7)
        ws.cell(row=3, column=7).value = '日期格式'
        # Data columns D/E/F were already correct; new G is empty for now
    else:
        # Already has both columns; ensure header names are correct
        ws.cell(row=3, column=7).value = '日期格式'
        ws.cell(row=3, column=8).value = '解析类型'
    wb.save(path)

# Main execution
for f in ['发票模版-评审/done/FBU-加拿大-5%GST-负数.xlsx',
          '发票模版-评审/done/FBU-加拿大-5%GST.xlsx']:
    wb = openpyxl.load_workbook(f)
    fix_canada_main(wb[wb.sheetnames[0]])
    wb.save(f)
    print(f'Fixed Canada main: {f}')

for f in ['发票模版-评审/done/FBU-美国EL发票模板-US credit note（负数）.xlsx',
          '发票模版-评审/done/FBU-美国GA发票模板-GAUS credit note（负.xlsx']:
    wb = openpyxl.load_workbook(f)
    fix_usa_main(wb[wb.sheetnames[0]])
    wb.save(f)
    print(f'Fixed USA main: {f}')

print('Done.')
