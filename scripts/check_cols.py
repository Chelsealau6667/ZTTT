import openpyxl
from openpyxl.utils import get_column_letter
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"
filepath = os.path.join(REVIEW_DIR, "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx")
wb = openpyxl.load_workbook(filepath)
ws = wb.worksheets[0]

# Check rows 28-36, all columns
for row in range(28, 37):
    row_data = []
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val:
            row_data.append(f"{get_column_letter(col)}='{val}'")
    print(f"Row {row}: {' | '.join(row_data)}")

print("\n--- Column widths ---")
for col in range(1, 9):
    letter = get_column_letter(col)
    width = ws.column_dimensions[letter].width
    print(f"Col {letter}: width={width}")
