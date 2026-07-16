import openpyxl
from openpyxl.utils import get_column_letter
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"
for fname in ["FBU-温哥华模版-GST税率(0%、5%）正数.xlsx", "FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx"]:
    filepath = os.path.join(REVIEW_DIR, fname)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    print(f"\nFILE: {fname}")
    for row in range(28, 40):
        row_data = []
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val:
                row_data.append(f"{get_column_letter(col)}='{val}'")
        if row_data:
            print(f"  Row {row}: {' | '.join(row_data)}")
