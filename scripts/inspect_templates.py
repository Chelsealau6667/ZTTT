import openpyxl
from openpyxl.styles import Font
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

def inspect_file(filepath, name):
    print(f"\n{'='*60}")
    print(f"FILE: {name}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Print all cells with their values and font colors in columns A-H
    for row in range(1, min(30, ws.max_row + 1)):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                font_color = None
                if cell.font and cell.font.color:
                    font_color = cell.font.color.rgb
                print(f"  [{row},{col}] '{cell.value}' color={font_color}")

# Inspect Canada HST positive
inspect_file(os.path.join(REVIEW_DIR, "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx"), "Canada HST Positive")
# Inspect Australia
inspect_file(os.path.join(REVIEW_DIR, "FBU-澳洲谷仓开票模板.xlsx"), "Australia")
# Inspect Japan
inspect_file(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"), "Japan")
# Inspect Shenzhen
inspect_file(os.path.join(REVIEW_DIR, "ABU-深圳云颂开票模板.xlsx"), "Shenzhen")
