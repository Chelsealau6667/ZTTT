import openpyxl
from openpyxl.styles import Font
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

def inspect_detail(filepath, name):
    print(f"\n{'='*60}")
    print(f"FILE: {name}")
    print(f"{'='*60}")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Print merged cell ranges
    merged = list(ws.merged_cells.ranges)
    if merged:
        print("Merged cells:")
        for m in merged:
            print(f"  {m}")
    
    # Focus on rows 27-35 for beneficiary block
    for row in range(27, min(36, ws.max_row + 1)):
        row_data = []
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                font_color = None
                if cell.font and cell.font.color:
                    font_color = cell.font.color.rgb
                row_data.append(f"[{col}]({font_color}):'{cell.value}'")
        if row_data:
            print(f"  Row {row}: {' | '.join(row_data)}")

# Canada HST Positive
inspect_detail(os.path.join(REVIEW_DIR, "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx"), "Canada HST Positive")
# Canada HST Negative
inspect_detail(os.path.join(REVIEW_DIR, "FBU-加拿大谷仓发票模版-HST税率13%（负数）.xlsx"), "Canada HST Negative")
# Australia
inspect_detail(os.path.join(REVIEW_DIR, "FBU-澳洲谷仓开票模板.xlsx"), "Australia")
# Japan
inspect_detail(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"), "Japan")
# Vancouver GST Positive
inspect_detail(os.path.join(REVIEW_DIR, "FBU-温哥华模版-GST税率(0%、5%）正数.xlsx"), "Vancouver GST Positive")
