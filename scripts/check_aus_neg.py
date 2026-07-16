import openpyxl
from openpyxl.utils import get_column_letter
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"
filepath = os.path.join(REVIEW_DIR, "FBU-澳洲谷仓开票模板.xlsx")
wb = openpyxl.load_workbook(filepath)
ws = wb.worksheets[0]
print("Australia full row 14-16:")
for row in range(13, 20):
    row_data = []
    for col in range(1, 9):
        cell = ws.cell(row=row, column=col)
        val = cell.value
        if val:
            color = None
            if cell.font and cell.font.color:
                color = str(cell.font.color.rgb)
            row_data.append(f"{get_column_letter(col)}({color})='{val}'")
    if row_data:
        print(f"  Row {row}: {' | '.join(row_data)}")
