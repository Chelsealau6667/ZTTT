import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

RED = "FFFF0000"
BLACK = "FF000000"

def set_red(cell):
    if isinstance(cell, MergedCell):
        return
    if cell.font:
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                        italic=cell.font.italic, color=RED)
    else:
        cell.font = Font(color=RED)

def set_black(cell):
    if isinstance(cell, MergedCell):
        return
    if cell.font:
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                        italic=cell.font.italic, color=BLACK)
    else:
        cell.font = Font(color=BLACK)

def clear_cell(cell):
    if isinstance(cell, MergedCell):
        return
    cell.value = None
    set_black(cell)
    cell.fill = PatternFill(fill_type=None)

def unmerge_and_clear_cols(ws, columns):
    """Unmerge any merged ranges in specified columns, then clear all cells."""
    # Find merged ranges that intersect with target columns
    ranges_to_unmerge = []
    for merged_range in list(ws.merged_cells.ranges):
        min_col = merged_range.min_col
        max_col = merged_range.max_col
        for target_col in columns:
            if min_col <= target_col <= max_col:
                ranges_to_unmerge.append(merged_range)
                break
    
    for mr in ranges_to_unmerge:
        ws.unmerge_cells(str(mr))
    
    # Now clear all cells in the target columns
    for row in range(1, ws.max_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            clear_cell(cell)

def fix_canada_vancouver(filepath, is_canada=True):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # 1. Fix TOTAL row
    total_row = 26 if is_canada else 28
    for col in range(3, 6):
        cell = ws.cell(row=total_row, column=col)
        if cell.value and "合计" in str(cell.value):
            set_red(cell)
    
    # 2. Fix beneficiary block
    if is_canada:
        beneficiary_rows = {
            29: ("Beneficiary:", "账户名称"),
            30: ("Beneficiary's address:", "销售方地址"),
            31: ("Bank name:", "银行名称"),
            32: ("Bank Address:", "银行地址"),
            33: ("Bank Transit:", "Bank Transit"),
            34: ("Fiancial Institution:", "Fiancial Institution"),
            35: ("SWIFT Code:", "银行联行码"),
            36: ("Account#:", "账户号（收款币种）"),
        }
    else:
        beneficiary_rows = {
            31: ("Beneficiary:", "账户名称"),
            32: ("Beneficiary's address:", "销售方地址"),
            33: ("Bank name:", "银行名称"),
            34: ("Bank Address:", "银行地址"),
            35: ("SWIFT Code:", "银行联行码"),
            36: ("Account#:", "账户号（收款币种）"),
        }
    
    for row, (label, value) in beneficiary_rows.items():
        label_cell = ws.cell(row=row, column=2)
        label_cell.value = label
        set_black(label_cell)
        
        value_cell = ws.cell(row=row, column=3)
        value_cell.value = value
        set_red(value_cell)
    
    # 3. Clear annotation columns F, G, H (cols 6,7,8)
    unmerge_and_clear_cols(ws, [6, 7, 8])
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_australia(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # 1. Fix line item: 行项目.税额 (col E, row 14)
    cell = ws.cell(row=14, column=5)
    if cell.value and "行项目.税额" in str(cell.value):
        set_red(cell)
    
    # 2. Fix beneficiary block
    beneficiary_rows = {
        24: ("Bank Name:", "银行名称"),
        25: ("Account Name:", "账户名称"),
        26: ("BSB:", "BSB"),
        27: ("Account No:", "账户号"),
    }
    
    for row, (label, value) in beneficiary_rows.items():
        label_cell = ws.cell(row=row, column=2)
        label_cell.value = label
        set_black(label_cell)
        
        value_cell = ws.cell(row=row, column=3)
        value_cell.value = value
        set_red(value_cell)
    
    # 3. Clear annotation columns G, H
    unmerge_and_clear_cols(ws, [7, 8])
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_japan(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Fix line item rows - ensure all 行项目 fields are red
    for row in range(20, 25):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_red(cell)
    
    unmerge_and_clear_cols(ws, [7, 8])
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_yunsong(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    for row in range(17, 30):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_red(cell)
    
    unmerge_and_clear_cols(ws, [7, 8])
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def main():
    canada_files = [
        "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx",
        "FBU-加拿大谷仓发票模版-HST税率13%（负数）.xlsx",
    ]
    for f in canada_files:
        fix_canada_vancouver(os.path.join(REVIEW_DIR, f), is_canada=True)
    
    vancouver_files = [
        "FBU-温哥华模版-GST税率(0%、5%）正数.xlsx",
        "FBU-温哥华模版-GST税率(0%、5%）负数.xlsx",
        "FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx",
        "FBU-温哥华模版-HST税率（13%、14%、15%）负数.xlsx",
    ]
    for f in vancouver_files:
        fix_canada_vancouver(os.path.join(REVIEW_DIR, f), is_canada=False)
    
    fix_australia(os.path.join(REVIEW_DIR, "FBU-澳洲谷仓开票模板.xlsx"))
    fix_japan(os.path.join(REVIEW_DIR, "FBU-日本谷仓开票模版.xlsx"))
    fix_yunsong(os.path.join(REVIEW_DIR, "ABU-深圳云颂开票模板.xlsx"))
    fix_yunsong(os.path.join(REVIEW_DIR, "ABU-香港云颂开票模板.xlsx"))

if __name__ == "__main__":
    main()
