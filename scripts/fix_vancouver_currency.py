import openpyxl
from openpyxl.styles import Font, Color

base = '发票模版-评审'
files = [
    'FBU-温哥华模版-GST税率(0%、5%）正数.xlsx',
    'FBU-温哥华模版-GST税率(0%、5%）负数.xlsx',
    'FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx',
    'FBU-温哥华模版-HST税率（13%、14%、15%）负数.xlsx',
]

for f in files:
    path = f'{base}/{f}'
    wb = openpyxl.load_workbook(path)
    sheet_name = wb.sheetnames[0]  # main template sheet
    ws = wb[sheet_name]
    
    # ===== Main Template Updates =====
    # Row 28 C/D/E: amount fields -> prepend 币种 with a space
    # Determine current values to handle both GST（无币种） and HST（有币种但无空格）
    c28 = ws.cell(row=28, column=3)
    d28 = ws.cell(row=28, column=4)
    e28 = ws.cell(row=28, column=5)
    
    # Replace: if starts with 币种 but no space, ensure space; if doesn't start with 币种, add it
    def ensure_currency_prefix(val):
        if not val:
            return val
        text = str(val).strip()
        if text.startswith('币种'):
            # ensure single space after 币种
            rest = text[2:].lstrip()
            return f'币种 {rest}'
        else:
            return f'币种 {text}'
    
    c28.value = ensure_currency_prefix(c28.value)
    d28.value = ensure_currency_prefix(d28.value)
    e28.value = ensure_currency_prefix(e28.value)
    
    # Row 38 C: Account# line — current: 账户号(收款币种) or 账户号（收款币种）
    c38 = ws.cell(row=38, column=3)
    if c38.value and isinstance(c38.value, str):
        # Replace 收款币种 with 币种 inside parentheses, keep full-width parentheses
        c38.value = c38.value.replace('（', '(').replace('）', ')')
        # do it in one pass for full-width brackets
        # Actually the original text might already be full-width
        # Let's just replace the literal substring
        c38.value = c38.value.replace('（收款币种）', '（币种）')
        c38.value = c38.value.replace('(收款币种)', '（币种）')
    
    # ===== Field Rules Updates =====
    ws_rule = wb['字段取值规则']
    # Insert a row after row 10 (当前 row 10 是 到期日)
    ws_rule.insert_rows(11)
    # Copy style from row 10 (到期日) to new row 11
    for col in range(1, ws_rule.max_column + 1):
        src = ws_rule.cell(row=10, column=col)
        dst = ws_rule.cell(row=11, column=col)
        if src.has_style:
            dst.font = src.font.copy()
            dst.border = src.border.copy()
            dst.fill = src.fill.copy()
            dst.number_format = src.number_format
            dst.protection = src.protection.copy()
            dst.alignment = src.alignment.copy()
    
    # Fill values for new row
    ws_rule.cell(row=11, column=2).value = '币种'
    ws_rule.cell(row=11, column=3).value = 'currency'
    ws_rule.cell(row=11, column=4).value = '申请单映射字段'
    ws_rule.cell(row=11, column=5).value = '字段映射'
    
    wb.save(path)
    print(f'Saved: {f}')

print('Done.')
