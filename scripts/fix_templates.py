import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import copy
import os

REVIEW_DIR = "/Users/a/Desktop/ZT/FSSC/R406/invoice/发票模版-评审"

# Color constants
RED = "FFFF0000"      # Standard red for placeholder values
RED_ALT = "FFF54A45"  # Alternative red
BLACK = "FF000000"    # Black for labels

def is_red_font(font):
    """Check if font color is red (either standard red)"""
    if font and font.color and font.color.rgb:
        rgb = str(font.color.rgb).upper()
        return rgb in (RED, RED_ALT, "FF0000", "F54A45", "FFFF0000", "FFF54A45")
    return False

def set_font_red(cell):
    """Set cell font to red, preserving other font properties"""
    if cell.font:
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                         italic=cell.font.italic, color=RED)
    else:
        cell.font = Font(color=RED)

def set_font_black(cell):
    """Set cell font to black, preserving other font properties"""
    if cell.font:
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold,
                         italic=cell.font.italic, color=BLACK)
    else:
        cell.font = Font(color=BLACK)

def is_placeholder_text(val):
    """Check if text is a placeholder field that should be red"""
    if not val:
        return False
    val = str(val).strip()
    # Fields that must be red (value fields)
    red_fields = [
        "账户名称", "Beneficiary", "销售方地址", "Beneficiary's address",
        "银行名称", "Bank name", "银行地址", "Bank Address",
        "Bank Transit", "Fiancial Institution", "银行联行码", "SWIFT Code",
        "账户号", "Account #", "AccountNo", "Account#", "Account No",
        "BSB", "Account No", "账户号",
        "客户名称", "客户地址", "购买方名称", "购买方地址",
        "行项目.数量", "行项目.单价", "行项目.金额", "行项目.税额", "行项目.含税金额",
        "行项目.物料描述", "行项目.发票行号",
    ]
    for field in red_fields:
        if field.lower() in val.lower():
            return True
    # Pattern matching for fields with parentheses containing Chinese
    if "(" in val and ")" in val:
        return True
    if ")" in val and any(c in val for c in ["收款", "Account"]):
        return True
    return False

def is_label_text(val):
    """Check if text is a label that should be black"""
    if not val:
        return False
    val = str(val).strip()
    labels = [
        "Beneficiary:", "Bank name:", "Bank Address:", "Please remit to",
        "Bank Transit:", "Fiancial Institution:", "SWIFT Code:", "Account#:",
        "BSB:", "Account No:",
    ]
    for label in labels:
        if val.startswith(label):
            return True
    return False

def fix_canada_vancouver_template(filepath):
    """Fix Canada and Vancouver templates"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]  # First sheet (template sheet)
    
    # Fix 1: "Please remit to" block - labels black, values red
    for row in range(1, ws.max_row + 1):
        for col in range(1, 6):  # A-E columns
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            
            # Handle "Please remit to" rows
            if "Please remit to" in val or "remit" in val.lower():
                # This label stays black
                set_font_black(cell)
                continue
            
            # Handle colon-separated labels like "Beneficiary: 账户名称"
            if ":" in val:
                parts = val.split(":", 1)
                label_part = parts[0].strip()
                value_part = parts[1].strip() if len(parts) > 1 else ""
                
                # The label before colon should be black
                # The value after colon should be red
                if is_label_text(val) or label_part in ["Beneficiary", "Bank name", "Bank Address", 
                                                          "Bank Transit", "Fiancial Institution", 
                                                          "SWIFT Code", "Account#", "BSB", "Account No"]:
                    set_font_black(cell)
                    # If the entire content is in one cell, we can't split it easily in openpyxl
                    # without using rich text (which is complex). For now, set the whole cell to black
                    # if it contains a label, but if it contains a value field, set to red.
                    if value_part and is_placeholder_text(value_part):
                        set_font_red(cell)
                    continue
            
            # Check for standalone placeholder values
            if is_placeholder_text(val):
                set_font_red(cell)
            
            # Check for standalone labels
            if is_label_text(val):
                set_font_black(cell)
    
    # Fix 2: Line item placeholders (row 17 typically) should be red
    for row in [17, 18, 19, 20, 21]:
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_font_red(cell)
    
    # Fix 3: Clear annotation columns (F, G, H) - set font black, no fill
    for row in range(1, ws.max_row + 1):
        for col in range(6, 9):  # F, G, H
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
            set_font_black(cell)
            cell.fill = PatternFill(fill_type=None)
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_australia_template(filepath):
    """Fix Australia template"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Fix B26 BSB: BSB and B27 Account No: 账户号
    for row in [26, 27]:
        cell = ws.cell(row=row, column=2)  # Column B
        if cell.value:
            val = str(cell.value).strip()
            if ":" in val:
                parts = val.split(":", 1)
                label = parts[0].strip()
                value = parts[1].strip() if len(parts) > 1 else ""
                
                # The value part should be red
                if is_placeholder_text(value) or value:
                    set_font_red(cell)
                else:
                    set_font_black(cell)
    
    # Fix line item placeholders
    for row in range(17, 25):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_font_red(cell)
    
    # Clear annotation columns
    for row in range(1, ws.max_row + 1):
        for col in range(7, 9):  # G, H
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
            set_font_black(cell)
            cell.fill = PatternFill(fill_type=None)
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_japan_template(filepath):
    """Fix Japan template"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Fix line item placeholders (rows 20-21)
    for row in range(20, 25):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_font_red(cell)
    
    # Clear annotation columns
    for row in range(1, ws.max_row + 1):
        for col in range(7, 9):  # G, H
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
            set_font_black(cell)
            cell.fill = PatternFill(fill_type=None)
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def fix_yunsong_template(filepath):
    """Fix Shenzhen/Hong Kong templates"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.worksheets[0]
    
    # Fix line item placeholders
    for row in range(17, 30):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.value and "行项目" in str(cell.value):
                set_font_red(cell)
    
    # Clear annotation columns
    for row in range(1, ws.max_row + 1):
        for col in [7, 8]:  # G, H
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.value = None
            set_font_black(cell)
            cell.fill = PatternFill(fill_type=None)
    
    wb.save(filepath)
    print(f"Fixed: {os.path.basename(filepath)}")

def main():
    # Process Canada templates
    canada_files = [
        "FBU-加拿大谷仓发票模版-HST税率13%（正数）.xlsx",
        "FBU-加拿大谷仓发票模版-HST税率13%（负数）.xlsx",
    ]
    
    # Process Vancouver templates
    vancouver_files = [
        "FBU-温哥华模版-GST税率(0%、5%）正数.xlsx",
        "FBU-温哥华模版-GST税率(0%、5%）负数.xlsx",
        "FBU-温哥华模版-HST税率(（13%、14%、15%）正数.xlsx",
        "FBU-温哥华模版-HST税率（13%、14%、15%）负数.xlsx",
    ]
    
    for filename in canada_files + vancouver_files:
        filepath = os.path.join(REVIEW_DIR, filename)
        if os.path.exists(filepath):
            fix_canada_vancouver_template(filepath)
    
    # Process Australia
    aus_file = "FBU-澳洲谷仓开票模板.xlsx"
    filepath = os.path.join(REVIEW_DIR, aus_file)
    if os.path.exists(filepath):
        fix_australia_template(filepath)
    
    # Process Japan
    japan_file = "FBU-日本谷仓开票模版.xlsx"
    filepath = os.path.join(REVIEW_DIR, japan_file)
    if os.path.exists(filepath):
        fix_japan_template(filepath)
    
    # Process Yunsong templates
    yunsong_files = [
        "ABU-深圳云颂开票模板.xlsx",
        "ABU-香港云颂开票模板.xlsx",
    ]
    for filename in yunsong_files:
        filepath = os.path.join(REVIEW_DIR, filename)
        if os.path.exists(filepath):
            fix_yunsong_template(filepath)

if __name__ == "__main__":
    main()
